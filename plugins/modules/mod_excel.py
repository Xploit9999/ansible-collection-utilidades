#!/usr/bin/python
# -*- coding: utf-8 -*-
DOCUMENTATION = r'''
---
module: mod_excel
short_description: Modifica archivos Excel escribiendo datos en rangos o usando delimitadores.
version_added: "1.0"
description:
  - Este módulo permite modificar archivos de Excel (.xlsx), escribiendo datos horizontalmente en una hoja específica.
  - Se puede usar un delimitador como referencia para insertar datos, o un rango de celdas definido por el usuario.
  - Si se usa el delimitador, este será reemplazado con los datos y movido a la siguiente fila.
  - Si se define un rango (`celda_inicial` y `celda_final`), se escriben los datos en la primera fila vacía disponible.
options:
  ruta:
    description:
      - Ruta absoluta al archivo de Excel a modificar.
    required: true
    type: str
  hoja:
    description:
      - Nombre de la hoja donde se desea escribir. Si se omite, se usará la hoja activa.
    required: false
    type: str
  delimitador:
    description:
      - Cadena que se usará como delimitador para encontrar la posición donde escribir los datos.
      - No se puede usar junto con C(celda_inicial) o C(celda_final).
    required: false
    type: str
  celda_inicial:
    description:
      - Celda inicial del rango (por ejemplo, A34). Requiere también C(celda_final).
    required: false
    type: str
  celda_final:
    description:
      - Celda final del rango (por ejemplo, J37). Requiere también C(celda_inicial).
    required: false
    type: str
  data:
    description:
      - Lista de listas o lista plana con los datos que se van a escribir.
      - Si se usa el delimitador, debe ser una lista plana. Si se usan rangos, se permite lista de listas.
    required: true
    type: list
author:
  - John (@Xploit999)
'''

EXAMPLES = r'''
# Usar delimitador
- name: Escribir valores usando delimitador
  mod_excel:
    ruta: "/ruta/al/archivo.xlsx"
    hoja: "Formulario"
    delimitador: "*/"
    data:
      - "valor1"
      - "valor2"
      - "valor3"

# Escribir una sola fila en un rango
- name: Escribir en rango definido
  mod_excel:
    ruta: "/ruta/al/archivo.xlsx"
    celda_inicial: "A34"
    celda_final: "J34"
    data:
      - ["dato1", "dato2", "dato3"]

# Escribir varias filas en el siguiente espacio vacío del rango
- name: Escribir múltiples filas
  mod_excel:
    ruta: "/ruta/al/archivo.xlsx"
    celda_inicial: "A34"
    celda_final: "J37"
    data:
      - ["dato1", "dato2", "dato3"]
      - ["dato4", "dato5", "dato6"]
      - ["dato7", "dato8", "dato9"]
'''
import openpyxl
from ansible.module_utils.basic import AnsibleModule

def busca_delimitador(sheet, delimitador):
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == delimitador:
                return cell
    return None

def celda_a_fila_columna(celda):
    col = ord(celda[0].upper()) - ord('A') + 1
    row = int(celda[1:])
    return row, col

def es_fila_vacia(sheet, fila, col_inicio, col_fin):
    for col in range(col_inicio, col_fin + 1):
        if sheet.cell(row=fila, column=col).value not in [None, ""]:
            return False
    return True

def escribir_datos_en_rango(sheet, celda_inicial, celda_final, data):
    fila_ini, col_ini = celda_a_fila_columna(celda_inicial)
    fila_fin, col_fin = celda_a_fila_columna(celda_final)
    
    datos_escritos = 0

    for fila in range(fila_ini, fila_fin + 1):
        if datos_escritos >= len(data):
            break

        if es_fila_vacia(sheet, fila, col_ini, col_fin):
            for offset, value in enumerate(data[datos_escritos]):
                col_actual = col_ini + offset
                if col_actual <= col_fin:
                    sheet.cell(row=fila, column=col_actual).value = value
            datos_escritos += 1

    return datos_escritos

def modificar_excel(modulo):
    ruta = modulo.params['ruta']
    hoja = modulo.params.get('hoja')
    delimitador = modulo.params.get('delimitador')
    data = modulo.params['data']
    celda_inicial = modulo.params.get('celda_inicial')
    celda_final = modulo.params.get('celda_final')

    if delimitador and (celda_inicial or celda_final):
        modulo.fail_json(msg="No se puede usar 'delimitador' con 'celda_inicial' o 'celda_final'.")

    try:
        workbook = openpyxl.load_workbook(ruta)
        sheet = workbook[hoja] if hoja else workbook.active

        if delimitador:
            celda = busca_delimitador(sheet, delimitador)
            if not celda:
                modulo.fail_json(msg=f"Delimitador '{delimitador}' no encontrado.")

            col = celda.column
            row = celda.row

            sheet.cell(row=row, column=col).value = None

            for i, val in enumerate(data):
                sheet.cell(row=row, column=col + i).value = val

            sheet.cell(row=row + 1, column=col).value = delimitador

        else:
            if not celda_inicial or not celda_final:
                modulo.fail_json(msg="Debe especificar 'celda_inicial' y 'celda_final' si no se usa 'delimitador'.")

            if not isinstance(data[0], list):
                data = [data]  

            filas_escritas = escribir_datos_en_rango(sheet, celda_inicial, celda_final, data)
            if filas_escritas == 0:
                modulo.fail_json(msg="No se encontraron filas vacías en el rango para escribir datos.")

        workbook.save(ruta)
        modulo.exit_json(changed=True, msg="El archivo se modificó correctamente.")

    except Exception as e:
        modulo.fail_json(msg=str(e))

def iniciar_proceso():
    argumentos = dict(
        ruta=dict(type='str', required=True),
        hoja=dict(type='str', required=False),
        delimitador=dict(type='str', required=False, default=None),
        data=dict(type='list', required=True),
        celda_inicial=dict(type='str', required=False, default=None),
        celda_final=dict(type='str', required=False, default=None)
    )

    modulo = AnsibleModule(
        argument_spec=argumentos,
        supports_check_mode=True
    )

    if modulo.check_mode:
        modulo.exit_json(changed=False)

    modificar_excel(modulo)

if __name__ == '__main__':
    iniciar_proceso()
